import PyPDF2
import docx
import json
import openpyxl

def read_file_content(file_path):
    """Read content from various file formats"""
    try:
        if file_path.endswith('.txt'):
            with open(file_path, 'r', encoding='utf-8') as f:
                return f.read()
        
        elif file_path.endswith('.pdf'):
            content = ""
            with open(file_path, 'rb') as f:
                pdf_reader = PyPDF2.PdfReader(f)
                for page in pdf_reader.pages:
                    content += page.extract_text() + "\n"
            return content
        
        elif file_path.endswith('.docx'):
            doc = docx.Document(file_path)
            content = "\n".join([paragraph.text for paragraph in doc.paragraphs])
            return content
        
        elif file_path.endswith('.json'):
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
                return json.dumps(data, indent=2)
        
        elif file_path.endswith('.xlsx'):
            content = ""
            workbook = openpyxl.load_workbook(file_path)
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                content += f"\n--- Sheet: {sheet_name} ---\n"
                
                for row in sheet.iter_rows(values_only=True):
                    row_content = " | ".join([str(cell) if cell is not None else "" for cell in row])
                    content += row_content + "\n"
            
            return content
        
        else:
            return ""
    
    except Exception as e:
        st.error(f"Error reading file {file_path}: {e}")
        return ""