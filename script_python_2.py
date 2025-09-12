import os
import re
from datetime import datetime
from ai21 import AI21Client
from ai21.models.chat import ChatMessage
from dotenv import load_dotenv
from pymongo import MongoClient
import chromadb
import uuid
import PyPDF2
import docx
import json
import shutil
from pathlib import Path
import openpyxl  

load_dotenv()
api_key = os.getenv("API_KEY")

if not api_key:
    print(" Error: API_KEY not found.")
    exit()

client = AI21Client(api_key=api_key)

try:
    mongo_client = MongoClient(os.getenv("MONGO_URI"))
    db = mongo_client["intern_data"]
    collection = db["chat_collection"]
    users_collection = db["users_collection"]
except Exception as e:
    print(" MongoDB Connection Error:", e)
    exit()

# ChromaDB Configuration
CHROMA_API_KEY = os.getenv("CHROMA_API_KEY")
CHROMA_TENANT = os.getenv("CHROMA_TENANT")
CHROMA_DB = os.getenv("CHROMA_DB")

# Initialize ChromaDB
try:
    if CHROMA_API_KEY:
        chroma_client = chromadb.CloudClient(
            api_key=CHROMA_API_KEY,
            database=CHROMA_DB
        )
        print("  ChromaDB Cloud initialized successfully! ")
    else:
        chroma_client = chromadb.PersistentClient(path="./chroma_db")
        print("  Local ChromaDB initialized! ")
except Exception as e:
    print("  ChromaDB Error:", e)
    print(" Falling back to local ChromaDB...")
    try:
        chroma_client = chromadb.PersistentClient(path="./chroma_db")
        print("  Local ChromaDB initialized as fallback! ")
    except Exception as e2:
        print("  Local ChromaDB also failed:", e2)
        exit()

# Create user-specific uploads directory
def get_user_uploads_dir(user_id):
    """Get user-specific uploads directory path"""
    user_dir = Path(f"./uploads/user_{user_id}")
    user_dir.mkdir(parents=True, exist_ok=True)
    return user_dir

def generate_user_id():
    last_user = users_collection.find_one(sort=[("created_at", -1)])
    if not last_user:
        return "USR0001"
    last_id = last_user["user_id"]
    last_num = int(last_id[3:])
    return f"USR{last_num+1:04d}"

def is_valid_email(email):
    return re.match(r"^[\w\.-]+@[\w\.-]+\.\w+$", email)

def register_user():
    username = input(" Enter your name: ").strip()
    while True:
        email = input(" Enter your email: ").strip().lower()
        if not is_valid_email(email):
            print(" Invalid email.")
            retry = input("Try again? (y/n): ").strip().lower()
            if retry == "y":
                continue
            else:
                return None, None
        if users_collection.find_one({"email": email}):
            print(" Email exists. Please login.")
            retry = input("Login? (y/n): ").strip().lower()
            if retry == "y":
                return login_user()
            else:
                continue
        break

    user_id = generate_user_id()
    users_collection.insert_one({
        "user_id": user_id,
        "username": username,
        "email": email,
        "created_at": datetime.now()
    })
    
    # Create user-specific upload directory
    user_uploads_dir = get_user_uploads_dir(user_id)
    print(f"  Registered successfully! User ID: {user_id}")
    return user_id, True

def login_user():
    while True:
        email = input(" Enter your email: ").strip().lower()
        user_id_input = input(" Enter your User ID: ").strip().upper()
        if not is_valid_email(email):
            print(" Invalid email.")
            retry = input("Try again? (y/n): ").strip().lower()
            if retry == "y":
                continue
            else:
                return None, None

        user = users_collection.find_one({"email": email, "user_id": user_id_input})
        if not user:
            print(" Not found. Try again.")
            retry = input("Try again? (y/n): ").strip().lower()
            if retry == "y":
                continue
            else:
                return None, None

        print(f" Welcome back, {user['username']}!")
        return user["user_id"], False

def choose_mode():
    while True:
        print("\n Choose chat mode:")
        print("1️⃣ Global Chat (AI Model) - Direct AI answers")
        print("2️⃣ Local Chat (Your Documents) - Answers from your files only")
        choice = input("Enter 1 or 2: ").strip()
        if choice == "1":
            return "global"
        elif choice == "2":
            return "local"
        else:
            print(" Invalid choice.")

def get_file_paths_from_uploads(user_id):
    """Get all file paths from user-specific uploads directory"""
    user_uploads_dir = get_user_uploads_dir(user_id)
    files = list(user_uploads_dir.glob("*"))
    return [str(f) for f in files if f.is_file()]

def upload_files(user_id):
    """Upload files to user-specific uploads directory"""
    print("\n File Upload")
    print("Drag and drop files here or enter full file paths")
    print("Supported formats: .txt, .pdf, .docx, .json, .xlsx")
    print("Press Enter twice to finish")
    
    uploaded_files = []
    user_uploads_dir = get_user_uploads_dir(user_id)
    
    while True:
        file_path = input("\nEnter file path: ").strip()
        
        if not file_path:
            if uploaded_files:
                break
            else:
                print(" Please add at least one file!")
                continue
            
        # Handle drag-and-drop paths 
        file_path = file_path.strip('"\'')
        
        if not os.path.exists(file_path):
            print(f"  File not found: {file_path}")
            continue
            
        # Check file extension
        allowed_extensions = {'.txt', '.pdf', '.docx', '.json', '.xlsx'}
        file_ext = os.path.splitext(file_path)[1].lower()
        
        if file_ext not in allowed_extensions:
            print(f"  Unsupported file type: {file_ext}")
            continue
            
        # Copy file to user-specific uploads directory
        filename = os.path.basename(file_path)
        dest_path = user_uploads_dir / filename
        
        # Handle duplicate files
        counter = 1
        while dest_path.exists():
            name, ext = os.path.splitext(filename)
            dest_path = user_uploads_dir / f"{name}_{counter}{ext}"
            counter += 1
            
        try:
            shutil.copy2(file_path, dest_path)
            uploaded_files.append(str(dest_path))
            print(f"  Uploaded: {filename}")
            
        except Exception as e:
            print(f"  Error uploading {filename}: {e}")
    
    return uploaded_files

def read_file_content(file_path):
    """Read content from various file formats"""
    try:
        if file_path.endswith('.txt'):
            with open(file_path, 'r', encoding='utf-8') as f:
                return f.read()
        
        elif file_path.endswith('.pdf'):
            content = ""
            with open(file_path, 'rb') as f:
                pdf_reader = PyPDF2.PdfReader(f)
                for page in pdf_reader.pages:
                    content += page.extract_text() + "\n"
            return content
        
        elif file_path.endswith('.docx'):
            doc = docx.Document(file_path)
            content = "\n".join([paragraph.text for paragraph in doc.paragraphs])
            return content
        
        elif file_path.endswith('.json'):
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
                return json.dumps(data, indent=2)
        
        elif file_path.endswith('.xlsx'):
            # Read Excel file
            content = ""
            workbook = openpyxl.load_workbook(file_path)
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                content += f"\n--- Sheet: {sheet_name} ---\n"
                
                for row in sheet.iter_rows(values_only=True):
                    row_content = " | ".join([str(cell) if cell is not None else "" for cell in row])
                    content += row_content + "\n"
            
            return content
        
        else:
            print(f" Unsupported file format: {file_path}")
            return ""
    
    except Exception as e:
        print(f" Error reading file {file_path}: {e}")
        return ""

def setup_document_collection(user_id):
    """Setup ChromaDB collection with documents for specific user"""
    print("\n Document Management")
    print("1. Upload new files")
    print("2. Use already uploaded files")
    print("3. View uploaded files")
    print("4. Clear all documents")
    print("5. Back to main menu")
    
    choice = input("Enter choice (1-5): ").strip()
    
    collection_name = f"user_{user_id}_documents"
    
    try:
        collection = chroma_client.get_or_create_collection(
            name=collection_name,
            metadata={"user_id": user_id, "created_at": datetime.now().isoformat()}
        )
        
        if choice == "1":
            uploaded_files = upload_files(user_id)
            if not uploaded_files:
                print(" No files uploaded!")
                return collection
            
            return process_files_to_collection(uploaded_files, collection, user_id)
        
        elif choice == "2":
            existing_files = get_file_paths_from_uploads(user_id)
            if not existing_files:
                print(" No files found in your uploads directory!")
                return collection
            
            print("\n Your uploaded files:")
            for i, file_path in enumerate(existing_files, 1):
                print(f"{i}. {os.path.basename(file_path)}")
            
            use_files = input("Use these files? (y/n): ").strip().lower()
            if use_files == 'y':
                return process_files_to_collection(existing_files, collection, user_id)
            else:
                return collection
        
        elif choice == "3":
            existing_files = get_file_paths_from_uploads(user_id)
            if not existing_files:
                print(" You haven't uploaded any files yet!")
            else:
                print("\n Your uploaded files:")
                for i, file_path in enumerate(existing_files, 1):
                    file_size = os.path.getsize(file_path) / 1024
                    print(f"{i}. {os.path.basename(file_path)} ({file_size:.1f} KB)")
            
            input("\nPress Enter to continue...")
            return collection
        
        elif choice == "4":
            confirm = input("Delete ALL your documents and uploaded files? (y/n): ").strip().lower()
            if confirm == 'y':
                chroma_client.delete_collection(collection_name)
                print("  All your documents cleared from database!")
                
                # Delete user-specific upload directory
                user_uploads_dir = get_user_uploads_dir(user_id)
                for file_path in user_uploads_dir.glob("*"):
                    if file_path.is_file():
                        file_path.unlink()
                print("  All your uploaded files deleted!")
                
                collection = chroma_client.get_or_create_collection(
                    name=collection_name,
                    metadata={"user_id": user_id, "cleared_at": datetime.now().isoformat()}
                )
            return collection
        
        elif choice == "5":
            return collection
        
        else:
            print(" Invalid choice!")
            return collection
    
    except Exception as e:
        print(f"  Error in document management: {e}")
        return None

def process_files_to_collection(file_paths, collection, user_id):
    """Process files and add to ChromaDB collection"""
    documents = []
    metadatas = []
    ids = []
    
    for file_path in file_paths:
        content = read_file_content(file_path)
        if content:
            sentences = re.split(r'(?<=[.!?]) +', content)
            chunks = []
            current_chunk = ""
            
            for sentence in sentences:
                if len(current_chunk) + len(sentence) < 800:
                    current_chunk += sentence + " "
                else:
                    chunks.append(current_chunk.strip())
                    current_chunk = sentence + " "
            
            if current_chunk:
                chunks.append(current_chunk.strip())
            
            for i, chunk in enumerate(chunks):
                documents.append(chunk)
                metadatas.append({
                    "source": file_path, 
                    "chunk_index": i,
                    "user_id": user_id,
                    "filename": os.path.basename(file_path),
                    "file_type": os.path.splitext(file_path)[1]
                })
                ids.append(f"{user_id}_{os.path.basename(file_path)}_{i}_{uuid.uuid4().hex[:6]}")
    
    if documents:
        collection.add(
            documents=documents,
            metadatas=metadatas,
            ids=ids
        )
        print(f"  Added {len(documents)} document chunks from {len(file_paths)} files!")
        print(f"  Total documents in your collection: {collection.count()}")
    else:
        print("  No valid content found in provided files!")
    
    return collection

def query_documents(question, collection, n_results=3):
    """Query documents using ChromaDB"""
    try:
        results = collection.query(
            query_texts=[question],
            n_results=n_results
        )
        
        if results and results['documents'] and results['documents'][0]:
            relevant_content = "\n\n".join([
                f" Source: {results['metadatas'][0][i].get('filename', 'Unknown')}\n"
                f"Content: {doc}\n"
                for i, doc in enumerate(results['documents'][0])
            ])
            return relevant_content
        else:
            return "  No relevant information found in your documents."
    
    except Exception as e:
        print(f"  Error querying documents: {e}")
        return " Error searching documents."

def list_my_chats(user_id):
    chats = list(collection.find({"user_id": user_id}).sort("timestamp", -1))
    return chats

def pick_chat_by_index(chats, prompt="Enter chat number: "):
    try:
        n = int(input(prompt).strip())
        if 1 <= n <= len(chats):
            return chats[n - 1]
        print(" Invalid choice.")
        return None
    except ValueError:
        print(" Enter a valid number.")
        return None

def chat_loop(conversation, chat_log, user_id, title=None, mode="global", doc_collection=None):
    print("\n  Type 'exit' to finish the chat.")
    
    while True:
        user_input = input("🧑 You: ").strip()
        if user_input.lower() in ["exit", "quit"]:
            print(" Saving chat...")
            break
        if user_input == "":
            print(" Please type something!")
            continue

        if mode == "local" and doc_collection:
            print("  Searching in your documents...")
            relevant_doc_content = query_documents(user_input, doc_collection)
            
            if relevant_doc_content and "No relevant information" not in relevant_doc_content:
                context_prompt = f"""Based ONLY on the following document content:

{relevant_doc_content}

Answer this question: {user_input}

If the document doesn't contain the exact answer, say "I don't have information about this in my documents". Do not use any external knowledge."""
                
                local_conversation = [ChatMessage(role="user", content=context_prompt)]
            else:
                local_conversation = [ChatMessage(role="user", content=user_input)]
            
            chat_log.append({"role": "user", "content": user_input})
            
            try:
                response = client.chat.completions.create(
                    messages=local_conversation,
                    model="jamba-large",
                    max_tokens=300,
                    temperature=0.1,
                    top_p=0.9
                )
                answer = response.choices[0].message.content
                print("🤖 Assistant (from your documents):", answer)
                conversation.append(ChatMessage(role="user", content=user_input))
                conversation.append(ChatMessage(role="assistant", content=answer))
                chat_log.append({"role": "assistant", "content": answer})
            except Exception as e:
                print(" AI21 API error:", e)
                if relevant_doc_content and "No relevant information" not in relevant_doc_content:
                    print(" Most relevant document content:", relevant_doc_content[:500] + "...")
                    conversation.append(ChatMessage(role="user", content=user_input))
                    conversation.append(ChatMessage(role="assistant", content=relevant_doc_content))
                    chat_log.append({"role": "assistant", "content": relevant_doc_content})
        
        else:
            conversation.append(ChatMessage(role="user", content=user_input))
            chat_log.append({"role": "user", "content": user_input})
            try:
                response = client.chat.completions.create(
                    messages=conversation,
                    model="jamba-large",
                    max_tokens=250,
                    temperature=0.3,
                    top_p=0.9
                )
                answer = response.choices[0].message.content
                print("🤖 Assistant:", answer)
                conversation.append(ChatMessage(role="assistant", content=answer))
                chat_log.append({"role": "assistant", "content": answer})
            except Exception as e:
                print(" AI21 API error:", e)

    if title is None:
        title = f"Chat_{datetime.now().strftime('%Y%m%d_%H%M%S')}"

    collection.insert_one({
        "user_id": user_id,
        "title": title,
        "chat": chat_log,
        "timestamp": datetime.now(),
        "mode": mode
    })
    print("  Chat saved to MongoDB!\n")

# ===== MAIN =====
print("  LOGIN / REGISTER")
if input("Do you have an account? (y/n): ").strip().lower() == "y":
    user_id, is_new = login_user()
    if not user_id:
        print(" Login cancelled.")
        exit()
else:
    user_id, is_new = register_user()
    if not user_id:
        print(" Registration cancelled.")
        exit()

mode = choose_mode()
doc_collection = None

if mode == "local":
    doc_collection = setup_document_collection(user_id)
    if not doc_collection:
        print(" Switching to Global mode due to document setup issues.")
        mode = "global"

if is_new:
    print("\n  Starting your first chat!")
    conversation = []
    chat_log = []
    chat_loop(conversation, chat_log, user_id, mode=mode, doc_collection=doc_collection)

# ===== MENU LOOP =====
while True:
    saved_chats = list_my_chats(user_id)

    if saved_chats:
        print("\n Menu:")
        print("1️⃣ Continue previous chat")
        print("2️⃣ Start new chat")
        print("3️⃣ Delete chat")
        if mode == "local":
            print("4️⃣ Manage documents")
        print("5️⃣ Exit")
        
        choice = input("Choose: ").strip()

        if choice == "1":
            for idx, c in enumerate(saved_chats, start=1):
                ts = c.get("timestamp")
                when = ts.strftime("%d %b %Y %H:%M") if ts else "-"
                print(f"{idx}. {c.get('title','Untitled')} — {when} ({c.get('mode', 'global')})")
            chosen = pick_chat_by_index(saved_chats, "Select chat: ")
            if not chosen:
                continue
            conversation = []
            chat_log = []
            print("\n Previous messages:")
            for msg in chosen.get("chat", []):
                emoji = "🧑" if msg['role']=="user" else "🤖"
                print(f"{emoji} {msg['content']}")
                conversation.append(ChatMessage(role=msg['role'], content=msg['content']))
                chat_log.append({"role": msg['role'], "content": msg['content']})
            
            chat_mode = chosen.get("mode", "global")
            current_doc_collection = doc_collection if chat_mode == "local" else None
            chat_loop(conversation, chat_log, user_id, title=chosen.get("title", "Untitled"), 
                     mode=chat_mode, doc_collection=current_doc_collection)

        elif choice == "2":
            conversation = []
            chat_log = []
            chat_loop(conversation, chat_log, user_id, mode=mode, doc_collection=doc_collection)

        elif choice == "3":
            for idx, c in enumerate(saved_chats, start=1):
                ts = c.get("timestamp")
                when = ts.strftime("%d %b %Y %H:%M") if ts else "-"
                print(f"{idx}. {c.get('title','Untitled')} — {when}")
            chosen = pick_chat_by_index(saved_chats, "Select chat to delete: ")
            if not chosen:
                continue
            confirm = input(f"Delete '{chosen.get('title','Untitled')}'? (y/n): ").strip().lower()
            if confirm == "y":
                collection.delete_one({"_id": chosen["_id"], "user_id": user_id})
                print("  Chat deleted.")
        
        elif choice == "4" and mode == "local":
            doc_collection = setup_document_collection(user_id)

        elif choice == "5":
            print("  Goodbye!")
            break
        else:
            print(" Invalid choice.")

    else:
        print("\n1️⃣ Start new chat")
        if mode == "local":
            print("2️⃣ Manage documents")
        print("3️⃣ Exit")
        
        choice = input("Choose: ").strip()
        if choice == "1":
            conversation = []
            chat_log = []
            chat_loop(conversation, chat_log, user_id, mode=mode, doc_collection=doc_collection)
        elif choice == "2" and mode == "local":
            doc_collection = setup_document_collection(user_id)
        elif choice == "3":
            print("  Goodbye!")
            break
        else:
            print(" Invalid choice.")